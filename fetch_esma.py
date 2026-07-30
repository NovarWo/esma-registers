#!/usr/bin/env python3
"""
ESMA MiCAR Register Tracker - scraper.

Downloads the 5 official ESMA "Interim MiCA Register" CSV files, normalises
each into a JSON file per register, diffs the result against the previous
snapshot committed in /data, and writes a changelog + meta file.

Designed to run twice a day via GitHub Actions. ESMA itself only republishes
the interim register on a weekly basis, so most runs will find "no change" -
that is expected, not a bug.

Source:
https://www.esma.europa.eu/esmas-activities/digital-finance-and-innovation/markets-crypto-assets-regulation-mica
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Iterable

import openpyxl
import requests

BASE_URL = "https://www.esma.europa.eu/sites/default/files/2024-12/{}.csv"

# register key (used for filenames / URLs in the site) -> ESMA CSV file stem
SOURCES = {
    "whitepapers": "OTHER",
    "art": "ARTZZ",
    "emt": "EMTWP",
    "casps": "CASPS",
    "non_compliant": "NCASP",
}

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
HISTORY_DIR = DATA_DIR / "history"

TIMEOUT = 30
USER_AGENT = "BTCDirect-ESMA-MiCAR-Tracker/1.0 (compliance monitoring; contact: compliance@btcdirect.eu)"


# --------------------------------------------------------------------------
# Fetch + raw CSV parsing
# --------------------------------------------------------------------------

def fetch_csv_text(register_code: str) -> str:
    url = BASE_URL.format(register_code)
    resp = requests.get(url, timeout=TIMEOUT, headers={"User-Agent": USER_AGENT})
    resp.raise_for_status()
    # ESMA files are UTF-8 with a leading BOM
    return resp.content.decode("utf-8-sig")


def parse_csv_text(text: str) -> list[dict]:
    """Parse ESMA's CSV text into a list of dicts, tolerant of their export quirks
    (trailing empty/unnamed columns, blank trailing rows, stray whitespace)."""
    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict] = []
    for raw_row in reader:
        clean = {}
        for key, value in raw_row.items():
            if not key or not key.strip():
                continue  # drop ESMA's trailing unnamed columns
            clean[key.strip()] = value.strip() if isinstance(value, str) else value
        if any(clean.values()):
            rows.append(clean)
    return rows


def fetch_csv(register_code: str) -> list[dict]:
    return parse_csv_text(fetch_csv_text(register_code))


# --------------------------------------------------------------------------
# Normalisation helpers
# --------------------------------------------------------------------------

def record_id(*parts) -> str:
    key = "|".join((p or "").strip() for p in parts)
    return hashlib.sha1(key.encode("utf-8")).hexdigest()[:16]


def split_pipe(value: str | None) -> list[str]:
    if not value:
        return []
    return [v.strip() for v in value.split("|") if v.strip()]


def status_from(end_date: str | None) -> str:
    return "withdrawn" if end_date else "active"


def group_rows(rows: Iterable[dict], key_fn: Callable[[dict], object],
               item_field: str, item_fn: Callable[[dict], dict | list[dict] | None]) -> list[dict]:
    """Group repeated ESMA rows (one row per service / whitepaper) into one
    record per entity, with the repeated bit collected into an array field.
    The first row seen for a key supplies the entity-level fields.

    `item_fn` may return a single dict, a list of dicts, or None/empty - some
    ESMA exports pack several values into one row (e.g. multiple services
    pipe-joined in a single `ac_serviceCode` cell) rather than repeating the
    row per value, so a single input row can expand into several items."""
    grouped: dict[object, dict] = {}
    order: list[object] = []
    for row in rows:
        key = key_fn(row)
        if key not in grouped:
            grouped[key] = {"_base": row, item_field: []}
            order.append(key)
        item = item_fn(row)
        if item:
            if isinstance(item, list):
                grouped[key][item_field].extend(item)
            else:
                grouped[key][item_field].append(item)
    result = []
    for key in order:
        bucket = grouped[key]
        record = dict(bucket["_base"])
        record[item_field] = bucket[item_field]
        result.append(record)
    return result


# --------------------------------------------------------------------------
# Per-register normalisers
# --------------------------------------------------------------------------

def normalize_casps(rows: list[dict]) -> list[dict]:
    def key_fn(r):
        return r.get("ae_lei") or (r.get("ae_lei_name", ""), r.get("ae_competentAuthority", ""))

    def item_fn(r):
        code = r.get("ac_serviceCode")
        if not code:
            return None
        # ESMA sometimes pipe-joins ALL of a CASP's services into this one
        # field on a single row, instead of repeating the row per service -
        # split defensively so each service becomes its own list item.
        countries = split_pipe(r.get("ac_serviceCode_cou"))
        comments = r.get("ac_comments") or None
        last_update = r.get("ac_lastupdate") or None
        return [
            {
                "service": part.strip(),
                "countries": countries,
                "comments": comments,
                "last_update": last_update,
            }
            for part in code.split("|") if part.strip()
        ]

    grouped = group_rows(rows, key_fn, "services", item_fn)
    records = []
    for r in grouped:
        end_date = r.get("ac_authorisationEndDate")
        records.append({
            "id": record_id("casps", r.get("ae_lei"), r.get("ae_lei_name"), r.get("ae_competentAuthority")),
            "competent_authority": r.get("ae_competentAuthority") or None,
            "home_member_state": r.get("ae_homeMemberState") or None,
            "name": r.get("ae_lei_name") or None,
            "lei": r.get("ae_lei") or None,
            "head_office_country": r.get("ae_lei_cou_code") or None,
            "commercial_name": r.get("ae_commercial_name") or None,
            "address": r.get("ae_address") or None,
            "website": r.get("ae_website") or None,
            "platform_website": r.get("ae_website_platform") or None,
            "authorisation_date": r.get("ac_authorisationNotificationDate") or None,
            "withdrawal_date": end_date or None,
            "status": status_from(end_date),
            "services": r["services"],
        })
    return records


def normalize_art_or_emt(rows: list[dict], register: str) -> list[dict]:
    is_emt = register == "emt"

    def key_fn(r):
        return r.get("ae_lei") or (r.get("ae_lei_name", ""), r.get("ae_competentAuthority", ""))

    def item_fn(r):
        url = r.get("wp_url")
        if not url:
            return None
        item = {
            "url": url,
            "start_date": r.get("wp_authorisationNotificationDate") or None,
            "dti": r.get("ae_DTI") or None,
            "dti_ffg": r.get("ae_DTI_FFG") or None,
            "comments": r.get("wp_comments") or None,
            "last_update": r.get("wp_lastupdate") or None,
        }
        if not is_emt:
            item["offer_countries"] = split_pipe(r.get("wp_url_cou"))
        return item

    grouped = group_rows(rows, key_fn, "whitepapers", item_fn)
    records = []
    for r in grouped:
        end_date = r.get("ac_authorisationEndDate")
        rec = {
            "id": record_id(register, r.get("ae_lei"), r.get("ae_lei_name"), r.get("ae_competentAuthority")),
            "competent_authority": r.get("ae_competentAuthority") or None,
            "home_member_state": r.get("ae_homeMemberState") or None,
            "name": r.get("ae_lei_name") or None,
            "lei": r.get("ae_lei") or None,
            "head_office_country": r.get("ae_lei_cou_code") or None,
            "commercial_name": r.get("ae_commercial_name") or None,
            "address": r.get("ae_address") or None,
            "website": r.get("ae_website") or None,
            "authorisation_date": r.get("ac_authorisationNotificationDate") or None,
            "withdrawal_date": end_date or None,
            "status": status_from(end_date),
            "whitepapers": r["whitepapers"],
        }
        if is_emt:
            rec["institution_type"] = r.get("ae_authorisation_other_emt") or None
            rec["exemption_48_4"] = r.get("ae_exemption48_4") or None
            rec["exemption_48_5"] = r.get("ae_exemption48_5") or None
        else:
            rec["credit_institution"] = r.get("ae_credit_institution") or None
        records.append(rec)
    return records


def normalize_whitepapers(rows: list[dict]) -> list[dict]:
    def key_fn(r):
        return r.get("ae_lei") or r.get("ae_lei_name", "")

    def item_fn(r):
        url = r.get("wp_url")
        if not url:
            return None
        return {
            "url": url,
            "casp_name": r.get("ae_lei_name_casp") or None,
            "casp_lei": r.get("ae_lei_casp") or None,
            "offer_countries": split_pipe(r.get("ae_offerCode_cou")),
            "dti": r.get("ae_DTI") or None,
            "dti_ffg": r.get("ae_DTI_FFG") or None,
            "comments": r.get("wp_comments") or None,
            "last_update": r.get("wp_lastupdate") or None,
        }

    grouped = group_rows(rows, key_fn, "whitepapers", item_fn)
    records = []
    for r in grouped:
        records.append({
            "id": record_id("whitepapers", r.get("ae_lei"), r.get("ae_lei_name")),
            "competent_authority": r.get("ae_competentAuthority") or None,
            "home_member_state": r.get("ae_homeMemberState") or None,
            "name": r.get("ae_lei_name") or None,
            "lei": r.get("ae_lei") or None,
            "head_office_country": r.get("ae_lei_cou_code") or None,
            "whitepapers": r["whitepapers"],
        })
    return records


def normalize_non_compliant(rows: list[dict]) -> list[dict]:
    seen: set[str] = set()
    records = []
    for r in rows:
        rid = record_id("ncasp", r.get("ae_lei"), r.get("ae_lei_name"),
                         r.get("ae_decision_date"), r.get("ae_website"))
        if rid in seen:
            continue  # exact duplicate row present in ESMA's own export
        seen.add(rid)
        records.append({
            "id": rid,
            "competent_authority": r.get("ae_competentAuthority") or None,
            "home_member_state": r.get("ae_homeMemberState") or None,
            "name": r.get("ae_lei_name") or None,
            "lei": r.get("ae_lei") or None,
            "head_office_country": r.get("ae_lei_cou_code") or None,
            "website": r.get("ae_website") or None,
            "article_17_infringement": r.get("ae_infrigment") or None,
            "reason": r.get("ae_reason") or None,
            "decision_date": r.get("ae_decision_date") or None,
            "comments": r.get("ae_comments") or None,
            "last_update": r.get("ae_lastupdate") or None,
        })
    return records


# --------------------------------------------------------------------------
# AFM's own crypto register (CASPs authorised/notified in the Netherlands).
#
# Unlike the 5 registers above (ESMA CSV exports), the AFM publishes theirs
# as a single .xlsx download - this is the register that matters most for a
# Dutch-licensed CASP, and AFM republishes it more often than ESMA's
# EU-wide consolidated register. Fetched and parsed separately, but feeds
# into the exact same diff_records()/Slack pipeline as the other 5 - see
# the "AFM register" block in run() below.
# --------------------------------------------------------------------------

AFM_XLSX_URL = "https://www.afm.nl/~/profmedia/files/registers/register-cryptopartijen.xlsx"

# AFM writes the home member state / passport countries as full English
# names / codes in free text rather than ESMA's consistent ISO codes -
# normalise to the same 2-letter codes used everywhere else on the site
# (assets/js/i18n.js's `countries` dict) so country flags/filters just work.
AFM_COUNTRY_NAME_TO_CODE = {
    "austria": "AT", "belgium": "BE", "bulgaria": "BG", "croatia": "HR",
    "cyprus": "CY", "czechia": "CZ", "czech republic": "CZ", "denmark": "DK",
    "estonia": "EE", "finland": "FI", "france": "FR", "germany": "DE",
    "greece": "EL", "hungary": "HU", "iceland": "IS", "ireland": "IE",
    "italy": "IT", "latvia": "LV", "liechtenstein": "LI", "lithuania": "LT",
    "luxembourg": "LU", "malta": "MT", "the netherlands": "NL", "netherlands": "NL",
    "norway": "NO", "poland": "PL", "portugal": "PT", "romania": "RO",
    "slovakia": "SK", "slovenia": "SI", "spain": "ES", "sweden": "SE",
    "united kingdom": "GB", "switzerland": "CH",
}
_AFM_KNOWN_EEA_CODES = set(AFM_COUNTRY_NAME_TO_CODE.values())
_AFM_SERVICE_LINE_RE = re.compile(r"^\(([a-j])\)\s*(.*)$", re.I)


def fetch_afm_rows() -> list[tuple]:
    """Downloads and parses the AFM crypto register (.xlsx). AFM's export has
    a few title/blank rows before the real header, and has reformatted this
    file before - rather than hardcoding a row offset, scan for the row
    whose first cell is literally "Entity name" and take everything after it."""
    resp = requests.get(AFM_XLSX_URL, timeout=TIMEOUT, headers={"User-Agent": USER_AGENT})
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    header_idx = next(
        (i for i, r in enumerate(all_rows) if r and isinstance(r[0], str) and r[0].strip() == "Entity name"),
        None,
    )
    if header_idx is None:
        raise ValueError("couldn't locate the header row (looked for a first cell reading 'Entity name')")
    return all_rows[header_idx + 1:]


def _afm_text(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.upper() == "N/A":
        return None
    return s


def _afm_date_str(value) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return _afm_text(value)


def _afm_auth_type(raw) -> str | None:
    text = _afm_text(raw)
    if not text:
        return None
    low = text.lower()
    if "notification" in low:
        return "notification"
    if "cross-border" in low or "cross border" in low:
        return "cross_border"
    if "authorisation" in low or "authorization" in low:
        return "authorisation"
    return "other"


def _parse_afm_services(raw) -> list[dict]:
    """AFM writes services as newline-separated "(x) description" lines
    (e.g. "(c) exchange of crypto-assets for funds"). Converted to the same
    "x. description" shape ESMA's CASPS register uses, with no per-service
    country breakdown (AFM only tracks passport countries at the entity
    level, see _parse_eu_passport) - this lets the front-end's existing
    CASP_SERVICES/extractServiceCode machinery render these with zero
    changes, since diff_records()'s service-level diffing also keys off
    this same leading-letter format."""
    text = _afm_text(raw)
    if not text:
        return []
    services = []
    for line in re.split(r"[\r\n]+", text):
        line = line.strip().rstrip(";").strip()
        if not line:
            continue
        m = _AFM_SERVICE_LINE_RE.match(line)
        if m:
            code, rest = m.group(1).lower(), m.group(2).strip()
            services.append({"service": f"{code}. {rest}", "countries": []})
        else:
            services.append({"service": line, "countries": []})
    return services


def _parse_websites(raw) -> tuple[str | None, str | None]:
    text = _afm_text(raw)
    if not text:
        return None, None
    parts = [p.strip() for p in re.split(r"[\r\n]+", text) if p.strip()]
    return (parts[0] if parts else None), (parts[1] if len(parts) > 1 else None)


def _parse_eu_passport(raw) -> tuple[str | None, list[str]]:
    text = _afm_text(raw)
    if not text:
        return None, []
    low = text.lower()
    direction = "outgoing" if low.startswith("outgoing") else "incoming" if low.startswith("incoming") else None
    codes: list[str] = []
    for tok in re.findall(r"\b[A-Za-z]{2}\b", text):
        up = "EL" if tok.upper() == "GR" else tok.upper()
        if up in _AFM_KNOWN_EEA_CODES and up not in codes:
            codes.append(up)
    return direction, codes


def normalize_afm(rows: list[tuple]) -> list[dict]:
    records = []
    for row in rows:
        if not row or not row[0] or not str(row[0]).strip():
            continue
        name = str(row[0]).strip()
        auth_number = _afm_text(row[1])
        home_state_raw = _afm_text(row[3])
        withdrawal_date = _afm_date_str(row[5])
        website, platform_website = _parse_websites(row[11] if len(row) > 11 else None)
        eu_passport_raw = _afm_text(row[12] if len(row) > 12 else None)
        eu_passport_direction, eu_passport_countries = _parse_eu_passport(row[12] if len(row) > 12 else None)

        records.append({
            "id": record_id("afm_casps", auth_number, name, home_state_raw),
            "name": name,
            "commercial_name": _afm_text(row[8] if len(row) > 8 else None),
            "lei": _afm_text(row[9] if len(row) > 9 else None),
            "authorisation_number": auth_number,
            "authorisation_type": _afm_auth_type(row[2] if len(row) > 2 else None),
            "home_member_state": AFM_COUNTRY_NAME_TO_CODE.get((home_state_raw or "").lower()),
            "authorisation_date": _afm_date_str(row[4] if len(row) > 4 else None),
            "withdrawal_date": withdrawal_date,
            "suspension_periods": _afm_text(row[6] if len(row) > 6 else None),
            "services": _parse_afm_services(row[7] if len(row) > 7 else None),
            "address": _afm_text(row[10] if len(row) > 10 else None),
            "website": website,
            "platform_website": platform_website,
            "eu_passport_direction": eu_passport_direction,
            "eu_passport_countries": eu_passport_countries,
            "eu_passport_raw": eu_passport_raw,
            "equivalent_services": _afm_text(row[13] if len(row) > 13 else None),
            "status": status_from(withdrawal_date),
        })
    return records


NORMALIZERS: dict[str, Callable[[list[dict]], list[dict]]] = {
    "whitepapers": normalize_whitepapers,
    "art": lambda rows: normalize_art_or_emt(rows, "art"),
    "emt": lambda rows: normalize_art_or_emt(rows, "emt"),
    "casps": normalize_casps,
    "non_compliant": normalize_non_compliant,
}


# --------------------------------------------------------------------------
# Diff + persistence
# --------------------------------------------------------------------------

def load_previous(register: str) -> dict[str, dict]:
    path = DATA_DIR / f"{register}.json"
    if not path.exists():
        return {}
    with path.open(encoding="utf-8") as f:
        data = json.load(f)
    return {rec["id"]: rec for rec in data.get("records", [])}


# Dutch short labels for the 10 MiCAR CASP services (Art. 3(1)(16), letters
# a-j) - mirrors assets/js/i18n.js's `services.<code>.label` (nl). Kept as a
# small duplicate here rather than shared with the front-end, since this
# script has no JS runtime available; if the i18n labels change, update both.
SERVICE_LABELS_NL = {
    "a": "Bewaring",
    "b": "Handelsplatform",
    "c": "Wisselen — fiat",
    "d": "Wisselen — crypto",
    "e": "Orderuitvoering",
    "f": "Plaatsing",
    "g": "Orderdoorgifte",
    "h": "Advies",
    "i": "Vermogensbeheer",
    "j": "Overdracht",
}

# Mirrors extractServiceCode() in assets/js/app.js: ESMA rows normally lead
# each service with its MiCAR letter code ("a. providing custody..."), but
# some real-world rows omit the letter - fall back to matching the official
# English service wording so those still resolve to a canonical service.
_SERVICE_CODE_RE = re.compile(r"^\s*([a-j])\s*[.)]", re.I)
_SERVICE_KEYWORDS = [
    ("a", re.compile(r"custody", re.I)),
    ("b", re.compile(r"trading platform", re.I)),
    ("c", re.compile(r"exchange of crypto-assets? for funds", re.I)),
    ("d", re.compile(r"exchange of crypto-assets? for other crypto", re.I)),
    ("e", re.compile(r"execution of orders", re.I)),
    ("f", re.compile(r"placing of crypto-assets", re.I)),
    ("g", re.compile(r"reception and transmission", re.I)),
    ("h", re.compile(r"advice on crypto-assets", re.I)),
    ("i", re.compile(r"portfolio management", re.I)),
    ("j", re.compile(r"transfer services", re.I)),
]


def extract_service_code(raw: str | None) -> str | None:
    if not raw:
        return None
    m = _SERVICE_CODE_RE.match(raw)
    if m:
        return m.group(1).lower()
    for code, pattern in _SERVICE_KEYWORDS:
        if pattern.search(raw):
            return code
    return None


def _index_services_by_code(services: list[dict] | None) -> dict[str, set[str]]:
    """code -> set of countries offered, merging pipe-joined/duplicate rows."""
    idx: dict[str, set[str]] = {}
    for s in services or []:
        for part in (s.get("service") or "").split("|"):
            code = extract_service_code(part.strip())
            if code:
                idx.setdefault(code, set()).update(s.get("countries") or [])
    return idx


def describe_service_changes(old_services: list[dict] | None, new_services: list[dict] | None) -> list[str]:
    """Human-readable (Dutch) lines describing what changed in a CASP's
    services: a new service type added, a service dropped, or - for a
    service offered in both snapshots - countries added/removed."""
    old_idx = _index_services_by_code(old_services)
    new_idx = _index_services_by_code(new_services)
    lines = []
    for code in sorted(new_idx.keys() - old_idx.keys()):
        lines.append(f"{SERVICE_LABELS_NL.get(code, code)} toegevoegd aan dienstverlening")
    for code in sorted(old_idx.keys() - new_idx.keys()):
        lines.append(f"{SERVICE_LABELS_NL.get(code, code)} niet langer aangeboden")
    for code in sorted(old_idx.keys() & new_idx.keys()):
        label = SERVICE_LABELS_NL.get(code, code)
        added = sorted(new_idx[code] - old_idx[code])
        removed = sorted(old_idx[code] - new_idx[code])
        if added:
            lines.append(f"{label} nu ook aangeboden in: {', '.join(added)}")
        if removed:
            lines.append(f"{label} niet langer aangeboden in: {', '.join(removed)}")
    return lines


def _fmt_value(v) -> str:
    if v is None or v == "":
        return "onbekend"
    return str(v)


def describe_record_change(old: dict, new: dict) -> list[str]:
    """Human-readable (Dutch) lines describing what changed between two
    snapshots of the same record - services get the detailed treatment above,
    every other field falls back to a generic "field: oud -> nieuw" line (or
    just "field gewijzigd" for list/dict-valued fields we don't want to dump
    raw into a Slack message, e.g. a register's "whitepapers" list)."""
    lines: list[str] = []
    if old.get("services") != new.get("services"):
        lines.extend(describe_service_changes(old.get("services"), new.get("services")))
    for key, new_val in new.items():
        if key in ("id", "services"):
            continue
        old_val = old.get(key)
        if old_val == new_val:
            continue
        label = key.replace("_", " ")
        if isinstance(old_val, (list, dict)) or isinstance(new_val, (list, dict)):
            lines.append(f"{label} gewijzigd")
        else:
            lines.append(f"{label}: {_fmt_value(old_val)} → {_fmt_value(new_val)}")
    return lines


def diff_records(register: str, previous: dict[str, dict], current: list[dict]) -> list[dict]:
    changes = []
    current_ids = set()
    for rec in current:
        current_ids.add(rec["id"])
        old = previous.get(rec["id"])
        if old is None:
            changes.append({"type": "added", "register": register, "id": rec["id"], "name": rec.get("name")})
        elif old != rec:
            entry = {"type": "changed", "register": register, "id": rec["id"], "name": rec.get("name")}
            detail = describe_record_change(old, rec)
            if detail:
                entry["detail"] = detail
            changes.append(entry)
    for old_id, old_rec in previous.items():
        if old_id not in current_ids:
            changes.append({"type": "removed", "register": register, "id": old_id, "name": old_rec.get("name")})
    return changes


def run(fetcher: Callable[[str], list[dict]] = fetch_csv) -> int:
    """Runs one full scrape cycle. `fetcher` is injectable for tests."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)

    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    all_changes: list[dict] = []
    counts: dict[str, int] = {}
    errors: dict[str, str] = {}

    for register, code in SOURCES.items():
        try:
            rows = fetcher(code)
        except Exception as exc:  # one register failing shouldn't kill the whole run
            errors[register] = str(exc)
            print(f"[warn] failed to fetch {register} ({code}): {exc}", file=sys.stderr)
            continue

        previous = load_previous(register)
        current = NORMALIZERS[register](rows)
        changes = diff_records(register, previous, current)
        all_changes.extend(changes)
        counts[register] = len(current)

        out_path = DATA_DIR / f"{register}.json"
        with out_path.open("w", encoding="utf-8") as f:
            json.dump(
                {"register": register, "generated_at": now, "records": current},
                f, ensure_ascii=False, indent=2,
            )

        if changes:
            print(f"[info] {register}: {len(changes)} change(s), {len(current)} total record(s)")
        else:
            print(f"[info] {register}: no change ({len(current)} records)")

    # AFM's own crypto register - fetched separately from the SOURCES loop
    # above since it's an .xlsx download rather than an ESMA CSV export (see
    # fetch_afm_rows()/normalize_afm()). Feeds the exact same
    # diff_records()/data-file/Slack pipeline as the other 5 registers.
    afm_register = "afm_casps"
    try:
        afm_rows = fetch_afm_rows()
        previous = load_previous(afm_register)
        current = normalize_afm(afm_rows)
        changes = diff_records(afm_register, previous, current)
        all_changes.extend(changes)
        counts[afm_register] = len(current)

        out_path = DATA_DIR / f"{afm_register}.json"
        with out_path.open("w", encoding="utf-8") as f:
            json.dump(
                {"register": afm_register, "generated_at": now, "records": current},
                f, ensure_ascii=False, indent=2,
            )

        if changes:
            print(f"[info] {afm_register}: {len(changes)} change(s), {len(current)} total record(s)")
        else:
            print(f"[info] {afm_register}: no change ({len(current)} records)")
    except Exception as exc:  # one register failing shouldn't kill the whole run
        errors[afm_register] = str(exc)
        print(f"[warn] failed to fetch {afm_register}: {exc}", file=sys.stderr)

    meta_path = DATA_DIR / "meta.json"
    meta = {}
    if meta_path.exists():
        with meta_path.open(encoding="utf-8") as f:
            meta = json.load(f)
    meta["last_checked"] = now
    meta["record_counts"] = counts
    if errors:
        meta["last_errors"] = errors
    else:
        meta.pop("last_errors", None)
    if all_changes:
        meta["last_change_at"] = now
    with meta_path.open("w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

    changelog_path = HISTORY_DIR / "changelog.json"
    changelog = []
    if changelog_path.exists():
        with changelog_path.open(encoding="utf-8") as f:
            changelog = json.load(f)
    for change in all_changes:
        changelog.append({**change, "timestamp": now})
    with changelog_path.open("w", encoding="utf-8") as f:
        json.dump(changelog, f, ensure_ascii=False, indent=2)

    print(f"[info] done - {len(all_changes)} total change(s) across all registers")

    # Every run rewrites each register's `generated_at` to "now", so a git
    # diff on data/ is never empty even when ESMA's actual content didn't
    # change - that's just freshness bookkeeping, not "new data". Expose the
    # real, record-level change count (and a human-readable per-entity
    # breakdown: which party, which register, added/changed/removed) as a
    # GitHub Actions step output, so the workflow can both gate its Slack
    # notification on genuinely new results instead of every successful run,
    # and say *what* changed rather than just how many.
    github_output = os.environ.get("GITHUB_OUTPUT")
    if github_output:
        TYPE_LABELS_NL = {"added": "toegevoegd", "changed": "gewijzigd", "removed": "verwijderd"}
        MAX_SUMMARY_LINES = 20
        detail_lines = []
        for c in all_changes:
            header = f"{c['register']}: {c.get('name') or c['id']}"
            # "changed" records with a field-level description (see
            # describe_record_change) get one line per changed aspect, e.g.
            # "casps: BTC Direct B.V. -> Bewaring toegevoegd aan dienstverlening"
            # - added/removed records, or a changed record where nothing
            # specific could be described, fall back to the plain type label.
            if c["type"] == "changed" and c.get("detail"):
                detail_lines.extend(f"{header} -> {d}" for d in c["detail"])
            else:
                detail_lines.append(f"{header} ({TYPE_LABELS_NL.get(c['type'], c['type'])})")
        summary_lines = detail_lines[:MAX_SUMMARY_LINES]
        if len(detail_lines) > MAX_SUMMARY_LINES:
            summary_lines.append(f"... en {len(detail_lines) - MAX_SUMMARY_LINES} andere wijziging(en)")
        with open(github_output, "a", encoding="utf-8") as f:
            f.write(f"real_changes={len(all_changes)}\n")
            f.write("change_summary<<EOF\n")
            f.write("\n".join(summary_lines) + "\n")
            f.write("EOF\n")

    if len(errors) == len(SOURCES) + 1:  # +1 for the AFM register, fetched separately above
        return 1  # every single register failed - a real failure, not "no change"
    return 0


def main() -> None:
    sys.exit(run())


if __name__ == "__main__":
    main()
