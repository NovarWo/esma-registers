#!/usr/bin/env python3
"""
ESMA MiCAR Register Tracker - scraper.

Downloads the 5 official ESMA "Interim MiCA Register" CSV files, normalises
each into a JSON file per register, diffs the result against the previous
snapshot committed in /data, and writes a changelog + meta file. The CASPs
register additionally folds in entities from AFM's own crypto register that
aren't in ESMA's export yet (see merge_esma_and_afm_casps) - AFM tends to
list a Dutch CASP's new service/authorisation before ESMA's slower, EU-wide
export catches up, so there's no separate "AFM register" on the site.

Designed to run hourly via GitHub Actions. ESMA itself only republishes the
interim register on a weekly basis, so most runs will find "no change" -
that is expected, not a bug.

Source:
https://www.esma.europa.eu/esmas-activities/digital-finance-and-innovation/markets-crypto-assets-regulation-mica
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Iterable

import openpyxl
import requests

BASE_URL = "https://www.esma.europa.eu/sites/default/files/2024-12/{}.csv"

# register key (used for filenames / URLs in the site) -> ESMA CSV file stem
SOURCES = {
    "whitepapers": "OTHER",
    "art": "ARTZZ",
    "emt": "EMTWP",
    "casps": "CASPS",
    "non_compliant": "NCASP",
}

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
HISTORY_DIR = DATA_DIR / "history"

TIMEOUT = 30
USER_AGENT = "BTCDirect-ESMA-MiCAR-Tracker/1.0 (compliance monitoring; contact: compliance@btcdirect.eu)"


# --------------------------------------------------------------------------
# Fetch + raw CSV parsing
# --------------------------------------------------------------------------

def fetch_csv_text(register_code: str) -> str:
    url = BASE_URL.format(register_code)
    resp = requests.get(url, timeout=TIMEOUT, headers={"User-Agent": USER_AGENT})
    resp.raise_for_status()
    # ESMA files are UTF-8 with a leading BOM
    return resp.content.decode("utf-8-sig")


def parse_csv_text(text: str) -> list[dict]:
    """Parse ESMA's CSV text into a list of dicts, tolerant of their export quirks
    (trailing empty/unnamed columns, blank trailing rows, stray whitespace)."""
    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict] = []
    for raw_row in reader:
        clean = {}
        for key, value in raw_row.items():
            if not key or not key.strip():
                continue  # drop ESMA's trailing unnamed columns
            clean[key.strip()] = value.strip() if isinstance(value, str) else value
        if any(clean.values()):
            rows.append(clean)
    return rows


def fetch_csv(register_code: str) -> list[dict]:
    return parse_csv_text(fetch_csv_text(register_code))


# --------------------------------------------------------------------------
# Normalisation helpers
# --------------------------------------------------------------------------

def record_id(*parts) -> str:
    key = "|".join((p or "").strip() for p in parts)
    return hashlib.sha1(key.encode("utf-8")).hexdigest()[:16]


def clean_lei(raw: str | None) -> str | None:
    """Normalises a raw LEI for use as both a display value and an identity
    key. Real-world ESMA/AFM exports occasionally carry a malformed LEI for
    the same entity between pulls - a stray trailing period, extra
    whitespace, or an off-by-one truncated character - which, left as-is,
    makes record_id() and _casps_match_keys()'s LEI match treat the SAME
    company as a different identity from one run to the next (observed in
    production: "Ramp Swaps (Ireland) Limited" with LEI
    "89450036UW3ID72T1M84." vs "89450036UW3ID72T1M84", and "WEB3 Technology
    B.V." with a 19-character LEI missing its last character - both caused
    duplicate rows and repeated spurious added/removed Slack notifications
    for the same real entity).

    Strips anything that isn't a letter or digit and uppercases (LEIs are
    always uppercase alphanumeric per ISO 17442), then validates the result
    is exactly 20 characters - the fixed length every real LEI has. Anything
    that doesn't clean up to that shape is treated as "no LEI" (record_id()/
    _casps_match_keys() then fall back to name-based identity/matching)
    rather than silently keeping a corrupt value that would never match
    anything anyway.

    Note: this can't fix every case - if a source genuinely replaces one
    valid-looking LEI with a different valid-looking one for the same entity
    (rather than just mangling the same LEI's formatting), that still reads
    as a LEI change; there's no way to tell that apart from an actual
    re-registration without also weighing the (also-fallible) entity name."""
    if not raw:
        return None
    cleaned = re.sub(r"[^A-Za-z0-9]", "", raw).upper()
    return cleaned if len(cleaned) == 20 else None


def split_pipe(value: str | None) -> list[str]:
    if not value:
        return []
    return [v.strip() for v in value.split("|") if v.strip()]


def status_from(end_date: str | None) -> str:
    return "withdrawn" if end_date else "active"


def group_rows(rows: Iterable[dict], key_fn: Callable[[dict], object],
               item_field: str, item_fn: Callable[[dict], dict | list[dict] | None]) -> list[dict]:
    """Group repeated ESMA rows (one row per service / whitepaper) into one
    record per entity, with the repeated bit collected into an array field.
    The first row seen for a key supplies the entity-level fields.

    `item_fn` may return a single dict, a list of dicts, or None/empty - some
    ESMA exports pack several values into one row (e.g. multiple services
    pipe-joined in a single `ac_serviceCode` cell) rather than repeating the
    row per value, so a single input row can expand into several items."""
    grouped: dict[object, dict] = {}
    order: list[object] = []
    for row in rows:
        key = key_fn(row)
        if key not in grouped:
            grouped[key] = {"_base": row, item_field: []}
            order.append(key)
        item = item_fn(row)
        if item:
            if isinstance(item, list):
                grouped[key][item_field].extend(item)
            else:
                grouped[key][item_field].append(item)
    result = []
    for key in order:
        bucket = grouped[key]
        record = dict(bucket["_base"])
        record[item_field] = bucket[item_field]
        result.append(record)
    return result


# --------------------------------------------------------------------------
# Per-register normalisers
# --------------------------------------------------------------------------

def normalize_casps(rows: list[dict]) -> list[dict]:
    def key_fn(r):
        return clean_lei(r.get("ae_lei")) or (r.get("ae_lei_name", ""), r.get("ae_competentAuthority", ""))

    def item_fn(r):
        code = r.get("ac_serviceCode")
        if not code:
            return None
        # ESMA sometimes pipe-joins ALL of a CASP's services into this one
        # field on a single row, instead of repeating the row per service -
        # split defensively so each service becomes its own list item.
        countries = split_pipe(r.get("ac_serviceCode_cou"))
        comments = r.get("ac_comments") or None
        last_update = r.get("ac_lastupdate") or None
        return [
            {
                "service": part.strip(),
                "countries": countries,
                "comments": comments,
                "last_update": last_update,
            }
            for part in code.split("|") if part.strip()
        ]

    grouped = group_rows(rows, key_fn, "services", item_fn)
    records = []
    for r in grouped:
        end_date = r.get("ac_authorisationEndDate")
        lei = clean_lei(r.get("ae_lei"))
        records.append({
            "id": record_id("casps", lei, r.get("ae_lei_name"), r.get("ae_competentAuthority")),
            "competent_authority": r.get("ae_competentAuthority") or None,
            "home_member_state": r.get("ae_homeMemberState") or None,
            "name": r.get("ae_lei_name") or None,
            "lei": lei,
            "head_office_country": r.get("ae_lei_cou_code") or None,
            "commercial_name": r.get("ae_commercial_name") or None,
            "address": r.get("ae_address") or None,
            "website": r.get("ae_website") or None,
            "platform_website": r.get("ae_website_platform") or None,
            "authorisation_date": r.get("ac_authorisationNotificationDate") or None,
            "withdrawal_date": end_date or None,
            "status": status_from(end_date),
            "services": r["services"],
        })
    return records


def normalize_art_or_emt(rows: list[dict], register: str) -> list[dict]:
    is_emt = register == "emt"

    def key_fn(r):
        return clean_lei(r.get("ae_lei")) or (r.get("ae_lei_name", ""), r.get("ae_competentAuthority", ""))

    def item_fn(r):
        url = r.get("wp_url")
        if not url:
            return None
        item = {
            "url": url,
            "start_date": r.get("wp_authorisationNotificationDate") or None,
            "dti": r.get("ae_DTI") or None,
            "dti_ffg": r.get("ae_DTI_FFG") or None,
            "comments": r.get("wp_comments") or None,
            "last_update": r.get("wp_lastupdate") or None,
        }
        if not is_emt:
            item["offer_countries"] = split_pipe(r.get("wp_url_cou"))
        return item

    grouped = group_rows(rows, key_fn, "whitepapers", item_fn)
    records = []
    for r in grouped:
        end_date = r.get("ac_authorisationEndDate")
        lei = clean_lei(r.get("ae_lei"))
        rec = {
            "id": record_id(register, lei, r.get("ae_lei_name"), r.get("ae_competentAuthority")),
            "competent_authority": r.get("ae_competentAuthority") or None,
            "home_member_state": r.get("ae_homeMemberState") or None,
            "name": r.get("ae_lei_name") or None,
            "lei": lei,
            "head_office_country": r.get("ae_lei_cou_code") or None,
            "commercial_name": r.get("ae_commercial_name") or None,
            "address": r.get("ae_address") or None,
            "website": r.get("ae_website") or None,
            "authorisation_date": r.get("ac_authorisationNotificationDate") or None,
            "withdrawal_date": end_date or None,
            "status": status_from(end_date),
            "whitepapers": r["whitepapers"],
        }
        if is_emt:
            rec["institution_type"] = r.get("ae_authorisation_other_emt") or None
            rec["exemption_48_4"] = r.get("ae_exemption48_4") or None
            rec["exemption_48_5"] = r.get("ae_exemption48_5") or None
        else:
            rec["credit_institution"] = r.get("ae_credit_institution") or None
        records.append(rec)
    return records


def normalize_whitepapers(rows: list[dict]) -> list[dict]:
    def key_fn(r):
        return clean_lei(r.get("ae_lei")) or r.get("ae_lei_name", "")

    def item_fn(r):
        url = r.get("wp_url")
        if not url:
            return None
        return {
            "url": url,
            "casp_name": r.get("ae_lei_name_casp") or None,
            "casp_lei": r.get("ae_lei_casp") or None,
            "offer_countries": split_pipe(r.get("ae_offerCode_cou")),
            "dti": r.get("ae_DTI") or None,
            "dti_ffg": r.get("ae_DTI_FFG") or None,
            "comments": r.get("wp_comments") or None,
            "last_update": r.get("wp_lastupdate") or None,
        }

    grouped = group_rows(rows, key_fn, "whitepapers", item_fn)
    records = []
    for r in grouped:
        lei = clean_lei(r.get("ae_lei"))
        records.append({
            "id": record_id("whitepapers", lei, r.get("ae_lei_name")),
            "competent_authority": r.get("ae_competentAuthority") or None,
            "home_member_state": r.get("ae_homeMemberState") or None,
            "name": r.get("ae_lei_name") or None,
            "lei": lei,
            "head_office_country": r.get("ae_lei_cou_code") or None,
            "whitepapers": r["whitepapers"],
        })
    return records


def normalize_non_compliant(rows: list[dict]) -> list[dict]:
    seen: set[str] = set()
    records = []
    for r in rows:
        lei = clean_lei(r.get("ae_lei"))
        rid = record_id("ncasp", lei, r.get("ae_lei_name"),
                         r.get("ae_decision_date"), r.get("ae_website"))
        if rid in seen:
            continue  # exact duplicate row present in ESMA's own export
        seen.add(rid)
        records.append({
            "id": rid,
            "competent_authority": r.get("ae_competentAuthority") or None,
            "home_member_state": r.get("ae_homeMemberState") or None,
            "name": r.get("ae_lei_name") or None,
            "lei": lei,
            "head_office_country": r.get("ae_lei_cou_code") or None,
            "website": r.get("ae_website") or None,
            "article_17_infringement": r.get("ae_infrigment") or None,
            "reason": r.get("ae_reason") or None,
            "decision_date": r.get("ae_decision_date") or None,
            "comments": r.get("ae_comments") or None,
            "last_update": r.get("ae_lastupdate") or None,
        })
    return records


# --------------------------------------------------------------------------
# AFM's own crypto register (CASPs authorised/notified in the Netherlands).
#
# Unlike the 5 registers above (ESMA CSV exports), the AFM publishes theirs
# as a single .xlsx download - this is the register that matters most for a
# Dutch-licensed CASP, and AFM republishes it more often than ESMA's
# EU-wide consolidated register. Fetched and parsed separately, but feeds
# into the exact same diff_records()/Slack pipeline as the other 5 - see
# the "AFM register" block in run() below.
# --------------------------------------------------------------------------

AFM_XLSX_URL = "https://www.afm.nl/~/profmedia/files/registers/register-cryptopartijen.xlsx"

# AFM writes the home member state / passport countries as full English
# names / codes in free text rather than ESMA's consistent ISO codes -
# normalise to the same 2-letter codes used everywhere else on the site
# (assets/js/i18n.js's `countries` dict) so country flags/filters just work.
AFM_COUNTRY_NAME_TO_CODE = {
    "austria": "AT", "belgium": "BE", "bulgaria": "BG", "croatia": "HR",
    "cyprus": "CY", "czechia": "CZ", "czech republic": "CZ", "denmark": "DK",
    "estonia": "EE", "finland": "FI", "france": "FR", "germany": "DE",
    "greece": "EL", "hungary": "HU", "iceland": "IS", "ireland": "IE",
    "italy": "IT", "latvia": "LV", "liechtenstein": "LI", "lithuania": "LT",
    "luxembourg": "LU", "malta": "MT", "the netherlands": "NL", "netherlands": "NL",
    "norway": "NO", "poland": "PL", "portugal": "PT", "romania": "RO",
    "slovakia": "SK", "slovenia": "SI", "spain": "ES", "sweden": "SE",
    "united kingdom": "GB", "switzerland": "CH",
}
_AFM_KNOWN_EEA_CODES = set(AFM_COUNTRY_NAME_TO_CODE.values())
_AFM_SERVICE_LINE_RE = re.compile(r"^\(([a-j])\)\s*(.*)$", re.I)


def fetch_afm_rows() -> list[tuple]:
    """Downloads and parses the AFM crypto register (.xlsx). AFM's export has
    a few title/blank rows before the real header, and has reformatted this
    file before - rather than hardcoding a row offset, scan for the row
    whose first cell is literally "Entity name" and take everything after it."""
    resp = requests.get(AFM_XLSX_URL, timeout=TIMEOUT, headers={"User-Agent": USER_AGENT})
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    header_idx = next(
        (i for i, r in enumerate(all_rows) if r and isinstance(r[0], str) and r[0].strip() == "Entity name"),
        None,
    )
    if header_idx is None:
        raise ValueError("couldn't locate the header row (looked for a first cell reading 'Entity name')")
    return all_rows[header_idx + 1:]


def _afm_text(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.upper() == "N/A":
        return None
    return s


def _afm_date_str(value) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return _afm_text(value)


def _afm_auth_type(raw) -> str | None:
    text = _afm_text(raw)
    if not text:
        return None
    low = text.lower()
    if "notification" in low:
        return "notification"
    if "cross-border" in low or "cross border" in low:
        return "cross_border"
    if "authorisation" in low or "authorization" in low:
        return "authorisation"
    return "other"


def _parse_afm_services(raw) -> list[dict]:
    """AFM writes services as newline-separated "(x) description" lines
    (e.g. "(c) exchange of crypto-assets for funds"). Converted to the same
    "x. description" shape ESMA's CASPS register uses, with no per-service
    country breakdown (AFM only tracks passport countries at the entity
    level, see _parse_eu_passport) - this lets the front-end's existing
    CASP_SERVICES/extractServiceCode machinery render these with zero
    changes, since diff_records()'s service-level diffing also keys off
    this same leading-letter format."""
    text = _afm_text(raw)
    if not text:
        return []
    services = []
    for line in re.split(r"[\r\n]+", text):
        line = line.strip().rstrip(";").strip()
        if not line:
            continue
        m = _AFM_SERVICE_LINE_RE.match(line)
        if m:
            code, rest = m.group(1).lower(), m.group(2).strip()
            services.append({"service": f"{code}. {rest}", "countries": []})
        else:
            services.append({"service": line, "countries": []})
    return services


def _parse_websites(raw) -> tuple[str | None, str | None]:
    text = _afm_text(raw)
    if not text:
        return None, None
    parts = [p.strip() for p in re.split(r"[\r\n]+", text) if p.strip()]
    return (parts[0] if parts else None), (parts[1] if len(parts) > 1 else None)


def _parse_eu_passport(raw) -> tuple[str | None, list[str]]:
    text = _afm_text(raw)
    if not text:
        return None, []
    low = text.lower()
    direction = "outgoing" if low.startswith("outgoing") else "incoming" if low.startswith("incoming") else None
    codes: list[str] = []
    for tok in re.findall(r"\b[A-Za-z]{2}\b", text):
        up = "EL" if tok.upper() == "GR" else tok.upper()
        if up in _AFM_KNOWN_EEA_CODES and up not in codes:
            codes.append(up)
    return direction, codes


def normalize_afm(rows: list[tuple]) -> list[dict]:
    records = []
    for row in rows:
        if not row or not row[0] or not str(row[0]).strip():
            continue
        name = str(row[0]).strip()
        auth_number = _afm_text(row[1])
        home_state_raw = _afm_text(row[3])
        withdrawal_date = _afm_date_str(row[5])
        website, platform_website = _parse_websites(row[11] if len(row) > 11 else None)
        eu_passport_raw = _afm_text(row[12] if len(row) > 12 else None)
        eu_passport_direction, eu_passport_countries = _parse_eu_passport(row[12] if len(row) > 12 else None)

        records.append({
            "id": record_id("afm_casps", auth_number, name, home_state_raw),
            "name": name,
            "commercial_name": _afm_text(row[8] if len(row) > 8 else None),
            "lei": clean_lei(_afm_text(row[9] if len(row) > 9 else None)),
            "authorisation_number": auth_number,
            "authorisation_type": _afm_auth_type(row[2] if len(row) > 2 else None),
            "home_member_state": AFM_COUNTRY_NAME_TO_CODE.get((home_state_raw or "").lower()),
            "authorisation_date": _afm_date_str(row[4] if len(row) > 4 else None),
            "withdrawal_date": withdrawal_date,
            "suspension_periods": _afm_text(row[6] if len(row) > 6 else None),
            "services": _parse_afm_services(row[7] if len(row) > 7 else None),
            "address": _afm_text(row[10] if len(row) > 10 else None),
            "website": website,
            "platform_website": platform_website,
            "eu_passport_direction": eu_passport_direction,
            "eu_passport_countries": eu_passport_countries,
            "eu_passport_raw": eu_passport_raw,
            "equivalent_services": _afm_text(row[13] if len(row) > 13 else None),
            "status": status_from(withdrawal_date),
        })
    return records


# --------------------------------------------------------------------------
# Merging AFM into the CASPs register.
#
# AFM's own crypto register only ever contains CASPs (there's no separate
# "AFM register" concept on the site) - AFM just tends to list a Dutch-
# licensed CASP's new service/authorisation *before* it shows up in ESMA's
# slower, EU-wide consolidated CASPS export. So instead of a 6th register,
# we fold AFM's entities straight into data/casps.json: an AFM entity is
# only added if it isn't already covered by ESMA's own export (matched by
# LEI, falling back to a normalised name match when a LEI is missing on
# either side) - once ESMA's own export catches up with the same entity,
# the AFM-derived stand-in is dropped and ESMA's version takes over, with
# no duplicate row and no spurious extra "added" notification (see the id
# continuity handling in merge_esma_and_afm_casps below).
# --------------------------------------------------------------------------

def _normalize_name(name: str | None) -> str:
    if not name:
        return ""
    n = name.lower().strip()
    n = re.sub(r"[^a-z0-9]+", " ", n)
    return re.sub(r"\s+", " ", n).strip()


def _casps_match_keys(rec: dict) -> list[tuple[str, str]]:
    """Every key that could recognise "the same real-world CASP" as this
    record, in priority order - not just "LEI if present, else name".

    A normalised-name key is always included alongside the LEI key (when
    there is one), because a single preferred-key-per-record scheme breaks
    the moment only ONE side of an ESMA/AFM pair has a usable LEI: that side
    matches on ("lei", ...), the other falls back to ("name", ...), and the
    two keys can never meet even though every _casps_match_keys() call is
    trying to recognise the exact same entity. This is exactly what happened
    with "WEB3 Technology B.V." in production - ESMA's LEI was truncated to
    19 characters (invalid per clean_lei()), AFM's was the real 20-character
    one, so the old single-key version put them in different buckets and
    the AFM row showed up as a spurious duplicate instead of being merged.
    Checking every key finds the match via name in that situation, while
    still preferring an actual LEI hit (listed first) when both records have
    one - LEI is the more reliable identifier when it's usable at all."""
    keys: list[tuple[str, str]] = []
    lei = clean_lei(rec.get("lei"))
    if lei:
        keys.append(("lei", lei))
    name = _normalize_name(rec.get("name"))
    if name:
        keys.append(("name", name))
    return keys


def _match_lookup(index: dict[tuple, object], keys: list[tuple[str, str]]):
    """Returns the first index hit across `keys`, in priority order, or None
    if none of them match - used to look a record up in a match-key index
    built by _casps_match_keys() without caring which specific key hit."""
    for key in keys:
        if key in index:
            return index[key]
    return None


def _map_afm_to_casps_shape(afm_rec: dict) -> dict:
    """Reshapes an AFM record into the same field shape ESMA's own CASPS
    records use, so the site's existing CASPs table/detail-view code needs
    no changes. AFM-only extras (authorisation number/type, EU-passport,
    equivalent services) are kept too - the front-end shows them in the
    detail view only for records tagged source == "afm"."""
    return {
        "competent_authority": "Netherlands Authority for the Financial Markets (AFM)",
        "home_member_state": afm_rec.get("home_member_state"),
        "name": afm_rec.get("name"),
        "lei": afm_rec.get("lei"),
        "head_office_country": None,
        "commercial_name": afm_rec.get("commercial_name"),
        "address": afm_rec.get("address"),
        "website": afm_rec.get("website"),
        "platform_website": afm_rec.get("platform_website"),
        "authorisation_date": afm_rec.get("authorisation_date"),
        "withdrawal_date": afm_rec.get("withdrawal_date"),
        "status": afm_rec.get("status"),
        "services": afm_rec.get("services"),
        "authorisation_number": afm_rec.get("authorisation_number"),
        "authorisation_type": afm_rec.get("authorisation_type"),
        "suspension_periods": afm_rec.get("suspension_periods"),
        "eu_passport_direction": afm_rec.get("eu_passport_direction"),
        "eu_passport_countries": afm_rec.get("eu_passport_countries"),
        "eu_passport_raw": afm_rec.get("eu_passport_raw"),
        "equivalent_services": afm_rec.get("equivalent_services"),
    }


def merge_esma_and_afm_casps(
    esma_current: list[dict], afm_current: list[dict], previous_merged: dict[str, dict]
) -> list[dict]:
    # previous_merged is {id: record} from load_previous("casps") - the last
    # committed, already-merged data/casps.json. Index it by EVERY match key
    # each record has (see _casps_match_keys()) so we can recognise "we've
    # tracked this exact entity before" regardless of which source
    # (co)produced it that time, and regardless of which key (LEI or name)
    # happens to be usable on either side this run.
    previous_by_key: dict[tuple, str] = {}
    for rec in previous_merged.values():
        for key in _casps_match_keys(rec):
            previous_by_key[key] = rec["id"]

    esma_by_key: dict[tuple, dict] = {}
    for rec in esma_current:
        for key in _casps_match_keys(rec):
            esma_by_key[key] = rec

    merged: list[dict] = []
    for rec in esma_current:
        keys = _casps_match_keys(rec)
        # If we've tracked this entity before (via ESMA or a since-superseded
        # AFM stand-in), keep its existing id rather than ESMA's freshly
        # computed one, so the tracked record's identity never changes.
        prev_id = _match_lookup(previous_by_key, keys)
        final_id = prev_id if prev_id is not None else rec["id"]
        merged.append({**rec, "id": final_id, "source": "esma"})

    for afm_rec in afm_current:
        keys = _casps_match_keys(afm_rec)
        if _match_lookup(esma_by_key, keys) is not None:
            continue  # already covered by ESMA's own export this run - skip, no duplicate row
        mapped = _map_afm_to_casps_shape(afm_rec)
        prev_id = _match_lookup(previous_by_key, keys)
        if prev_id is not None:
            mapped["id"] = prev_id
        elif keys:
            mapped["id"] = record_id("casps_afm", keys[0][1])
        else:
            mapped["id"] = afm_rec["id"]
        merged.append({**mapped, "source": "afm"})

    return merged


NORMALIZERS: dict[str, Callable[[list[dict]], list[dict]]] = {
    "whitepapers": normalize_whitepapers,
    "art": lambda rows: normalize_art_or_emt(rows, "art"),
    "emt": lambda rows: normalize_art_or_emt(rows, "emt"),
    "casps": normalize_casps,
    "non_compliant": normalize_non_compliant,
}


# --------------------------------------------------------------------------
# Diff + persistence
# --------------------------------------------------------------------------

def load_previous(register: str) -> dict[str, dict]:
    path = DATA_DIR / f"{register}.json"
    if not path.exists():
        return {}
    with path.open(encoding="utf-8") as f:
        data = json.load(f)
    return {rec["id"]: rec for rec in data.get("records", [])}


# Dutch short labels for the 10 MiCAR CASP services (Art. 3(1)(16), letters
# a-j) - mirrors assets/js/i18n.js's `services.<code>.label` (nl). Kept as a
# small duplicate here rather than shared with the front-end, since this
# script has no JS runtime available; if the i18n labels change, update both.
SERVICE_LABELS_NL = {
    "a": "Bewaring",
    "b": "Handelsplatform",
    "c": "Wisselen — fiat",
    "d": "Wisselen — crypto",
    "e": "Orderuitvoering",
    "f": "Plaatsing",
    "g": "Orderdoorgifte",
    "h": "Advies",
    "i": "Vermogensbeheer",
    "j": "Overdracht",
}

# Mirrors extractServiceCode() in assets/js/app.js: ESMA rows normally lead
# each service with its MiCAR letter code ("a. providing custody..."), but
# some real-world rows omit the letter - fall back to matching the official
# English service wording so those still resolve to a canonical service.
_SERVICE_CODE_RE = re.compile(r"^\s*([a-j])\s*[.)]", re.I)
_SERVICE_KEYWORDS = [
    ("a", re.compile(r"custody", re.I)),
    ("b", re.compile(r"trading platform", re.I)),
    ("c", re.compile(r"exchange of crypto-assets? for funds", re.I)),
    ("d", re.compile(r"exchange of crypto-assets? for other crypto", re.I)),
    ("e", re.compile(r"execution of orders", re.I)),
    ("f", re.compile(r"placing of crypto-assets", re.I)),
    ("g", re.compile(r"reception and transmission", re.I)),
    ("h", re.compile(r"advice on crypto-assets", re.I)),
    ("i", re.compile(r"portfolio management", re.I)),
    ("j", re.compile(r"transfer services", re.I)),
]


def extract_service_codes(raw: str | None) -> list[str]:
    """Most ESMA rows name exactly one service per segment (after splitting on
    "|"), usually with a leading letter prefix ("a. providing custody...") -
    for those, the prefix is authoritative and wins outright, no need to also
    keyword-scan the rest of the sentence. But some real-world rows (e.g.
    "Ronin EM Ltd") write ALL of a CASP's services as a single unprefixed,
    "/"-joined sentence instead of repeating/pipe-joining per service - a
    plain first-match keyword scan would then silently recognise only
    whichever service happened to be mentioned first and drop the rest. So
    without a prefix, this scans for every keyword that appears anywhere in
    the text and returns all of them, not just the first."""
    if not raw:
        return []
    m = _SERVICE_CODE_RE.match(raw)
    if m:
        return [m.group(1).lower()]
    return [code for code, pattern in _SERVICE_KEYWORDS if pattern.search(raw)]


def extract_service_code(raw: str | None) -> str | None:
    """Back-compat single-code accessor - prefer extract_service_codes()."""
    codes = extract_service_codes(raw)
    return codes[0] if codes else None


def _index_services_by_code(services: list[dict] | None) -> dict[str, set[str]]:
    """code -> set of countries offered, merging pipe-joined/duplicate rows."""
    idx: dict[str, set[str]] = {}
    for s in services or []:
        for part in (s.get("service") or "").split("|"):
            for code in extract_service_codes(part.strip()):
                idx.setdefault(code, set()).update(s.get("countries") or [])
    return idx


def describe_service_changes(old_services: list[dict] | None, new_services: list[dict] | None) -> list[dict]:
    """Structured, language-neutral description of what changed in a CASP's
    services: a new service type added, a service dropped, or - for a service
    offered in both snapshots - countries added/removed.

    This used to return pre-formatted Dutch sentences directly, which baked a
    fixed language into data/history/changelog.json - fine for the (always
    Dutch) Slack notification, but wrong once the *website* displays the same
    "what changed" text: a visitor with the EN toggle selected would still see
    Dutch service names like "Wisselen — fiat" inside an otherwise-English
    detail panel. Returning small {kind, code, ...} dicts instead lets each
    consumer render its own language: format_change_line_nl() below expands
    them to the Dutch sentences for Slack, and describeChangeLine() in
    assets/js/app.js expands them via the site's own t() for the viewer's
    current language."""
    old_idx = _index_services_by_code(old_services)
    new_idx = _index_services_by_code(new_services)
    lines: list[dict] = []
    for code in sorted(new_idx.keys() - old_idx.keys()):
        lines.append({"kind": "service_added", "code": code})
    for code in sorted(old_idx.keys() - new_idx.keys()):
        lines.append({"kind": "service_removed", "code": code})
    for code in sorted(old_idx.keys() & new_idx.keys()):
        added = sorted(new_idx[code] - old_idx[code])
        removed = sorted(old_idx[code] - new_idx[code])
        if added:
            lines.append({"kind": "service_countries_added", "code": code, "countries": added})
        if removed:
            lines.append({"kind": "service_countries_removed", "code": code, "countries": removed})
    return lines


def format_change_line_nl(line: dict | str) -> str:
    """Renders one describe_record_change() item as the Dutch sentence it used
    to be, for the (always-Dutch) Slack notification. Generic field-diff
    strings (e.g. "status: active -> withdrawn") pass through unchanged;
    structured dicts get expanded - service-change ones via SERVICE_LABELS_NL,
    "field_changed" via its own field name."""
    if isinstance(line, str):
        return line
    kind = line["kind"]
    if kind == "field_changed":
        return f"{line['field'].replace('_', ' ')} gewijzigd"
    label = SERVICE_LABELS_NL.get(line["code"], line["code"])
    if kind == "service_added":
        return f"{label} toegevoegd aan dienstverlening"
    if kind == "service_removed":
        return f"{label} niet langer aangeboden"
    if kind == "service_countries_added":
        return f"{label} nu ook aangeboden in: {', '.join(line['countries'])}"
    if kind == "service_countries_removed":
        return f"{label} niet langer aangeboden in: {', '.join(line['countries'])}"
    return label


def _format_countries(countries: tuple[str, ...], max_named: int) -> str:
    """Spells out a country list up to `max_named` codes, then folds the rest
    into a "+N andere" remainder - without this, a service gaining EU-passport
    coverage in, say, 20 countries at once would print all 20 codes on a
    single Slack line, exactly the kind of clutter this whole redesign is
    meant to avoid."""
    if len(countries) <= max_named:
        return ", ".join(countries)
    shown = countries[:max_named]
    return f"{', '.join(shown)} +{len(countries) - max_named} andere"


def summarize_change_detail(detail: list[dict | str]) -> str:
    """Combines every detail item for a single changed record into ONE Slack
    line, instead of the old one-line-per-item behaviour that made a record
    with many detail items (e.g. a CASP adding 7 services in the same new
    country) repeat its own name on 7 separate, near-identical lines.

    Service-change dicts that share the same kind and country list are
    grouped into a single segment (e.g. "Bewaring, Wisselen — fiat nu ook
    aangeboden in: DK" instead of two separate lines), and three independent
    caps keep any one record from blowing up the whole message:
    - MAX_SERVICES_NAMED: past this many services in one group, name the
      count instead of every service ("7 diensten ..." instead of listing
      all 7).
    - MAX_COUNTRIES_NAMED: past this many countries in one list, name the
      first few plus a "+N andere" remainder (see _format_countries()).
    - MAX_SEGMENTS: past this many distinct aspects changed on one record
      (mixing service changes with generic field diffs, say), name the first
      few plus a "+N andere wijziging(en)" remainder.

    Generic field-diff strings (e.g. "status: actief → ingetrokken") pass
    through as their own segment, appended after the grouped service
    segments - see describe_record_change(). "field_changed" dicts (a list/
    dict-valued field like "whitepapers" that changed shape - too complex to
    diff into a readable sentence) render the same way via
    format_change_line_nl(), since - unlike the service kinds - there's
    nothing to group them by (no code/countries in common)."""
    MAX_SERVICES_NAMED = 4
    MAX_COUNTRIES_NAMED = 6
    MAX_SEGMENTS = 3

    groups: dict[tuple, list[str]] = {}
    group_order: list[tuple] = []
    strings: list[str] = []

    for item in detail:
        if isinstance(item, str):
            strings.append(item)
            continue
        if item["kind"] == "field_changed":
            strings.append(format_change_line_nl(item))
            continue
        key = (item["kind"], tuple(item.get("countries", [])))
        if key not in groups:
            groups[key] = []
            group_order.append(key)
        groups[key].append(SERVICE_LABELS_NL.get(item["code"], item["code"]))

    segments: list[str] = []
    for kind, countries in group_order:
        labels = groups[(kind, countries)]
        name_part = f"{len(labels)} diensten" if len(labels) > MAX_SERVICES_NAMED else ", ".join(labels)
        if kind == "service_added":
            segments.append(f"{name_part} toegevoegd aan dienstverlening")
        elif kind == "service_removed":
            segments.append(f"{name_part} niet langer aangeboden")
        elif kind == "service_countries_added":
            segments.append(f"{name_part} nu ook aangeboden in: {_format_countries(countries, MAX_COUNTRIES_NAMED)}")
        elif kind == "service_countries_removed":
            segments.append(f"{name_part} niet langer aangeboden in: {_format_countries(countries, MAX_COUNTRIES_NAMED)}")
        else:
            segments.append(name_part)
    segments.extend(strings)

    if len(segments) > MAX_SEGMENTS:
        segments = segments[:MAX_SEGMENTS] + [f"+{len(segments) - MAX_SEGMENTS} andere wijziging(en)"]

    return "; ".join(segments)


def _fmt_value(v) -> str:
    if v is None or v == "":
        return "onbekend"
    return str(v)


def describe_record_change(old: dict, new: dict) -> list[dict | str]:
    """Lines describing what changed between two snapshots of the same
    record - services get the structured (language-neutral) treatment above;
    a plain scalar field ("field: oud -> nieuw") is a language-invariant
    string, since it's just the raw field name plus raw values (fine to pass
    through unchanged in both Slack and the website, on either language - see
    format_change_line_nl() and describeChangeLine() in assets/js/app.js).

    A list/dict-valued field (e.g. a register's "whitepapers" list) is too
    complex to diff into a readable "oud -> nieuw" string, so it gets a
    generic {"kind": "field_changed", "field": key} marker instead of a
    pre-baked "field gewijzigd" string - unlike the scalar case, that word
    itself ("gewijzigd"/"changed") IS language-dependent, so it needs the same
    structured treatment as the service-change kinds above to render correctly
    on an English-language site visit."""
    lines: list[dict | str] = []
    if old.get("services") != new.get("services"):
        lines.extend(describe_service_changes(old.get("services"), new.get("services")))
    for key, new_val in new.items():
        if key in ("id", "services", "source"):
            # "source" is internal bookkeeping (which upstream feed - ESMA or
            # AFM - most recently supplied this record, see
            # merge_esma_and_afm_casps), not a regulatory fact, so it's never
            # worth a line of its own in a Slack message.
            continue
        old_val = old.get(key)
        if old_val == new_val:
            continue
        if isinstance(old_val, (list, dict)) or isinstance(new_val, (list, dict)):
            lines.append({"kind": "field_changed", "field": key})
        else:
            label = key.replace("_", " ")
            lines.append(f"{label}: {_fmt_value(old_val)} → {_fmt_value(new_val)}")
    return lines


def _comparable(rec: dict) -> dict:
    """Strips fields that shouldn't by themselves make two snapshots of a
    record count as "changed" - currently just "source" (internal bookkeeping
    - which upstream feed, ESMA or AFM, most recently supplied this CASP's
    data, see merge_esma_and_afm_casps). Without this, an AFM-sourced CASP
    "graduating" into ESMA's own export - with every real field identical -
    would still register as a changed record (source: afm -> esma) and fire a
    Slack notification about nothing of substance."""
    return {k: v for k, v in rec.items() if k != "source"}


# User-requested importance ranking for the Slack notification's line-by-line
# summary: CASPs matter most (a brand-new CASP most of all - it's floated
# above every other CASP change too, not just other registers), then EMT,
# ART, Whitepapers, and Non-compliant last. This only reorders how the Slack
# summary is written - data/history/changelog.json itself still gets each
# run's changes appended in plain per-register order; the site applies the
# equivalent ranking when *displaying* the changelog instead (see
# changelogPriorityKey()/sortChangelogForDisplay() in assets/js/app.js).
REGISTER_PRIORITY = {"casps": 0, "emt": 1, "art": 2, "whitepapers": 3, "non_compliant": 4}

# Short display labels for the Slack summary (see run()'s github_output block
# below) - mirrors the site's nav.* labels closely enough for a one-word
# register tag, but isn't shared with i18n.js since Slack messages are always
# Dutch regardless of a visitor's site language preference.
REGISTER_LABELS_NL = {
    "casps": "CASPs",
    "emt": "EMT",
    "art": "ART",
    "whitepapers": "Whitepapers",
    "non_compliant": "Non-compliant",
}


def _change_priority_key(c: dict) -> tuple:
    register_rank = REGISTER_PRIORITY.get(c["register"], 99)
    is_new_casp = 0 if (c["register"] == "casps" and c["type"] == "added") else 1
    return (register_rank, is_new_casp)


# Once a record is removed, data/<register>.json no longer has it - there's
# nothing left to look up its original registration date from. So a
# "removed" changelog entry snapshots it from the last-known record right
# here, at the moment of removal, into the entry itself (see the site's
# openRemovedDetail() in assets/js/app.js, which shows this alongside the
# entry's own timestamp as the removal date). Non-compliant has no
# authorisation date (it was never authorised) - its decision_date is the
# closest equivalent "when this was first recorded" fact. Whitepapers has no
# suitable entity-level date field at all, so it's intentionally omitted.
REGISTRATION_DATE_FIELD = {
    "casps": "authorisation_date",
    "art": "authorisation_date",
    "emt": "authorisation_date",
    "non_compliant": "decision_date",
}


def diff_records(register: str, previous: dict[str, dict], current: list[dict]) -> list[dict]:
    changes = []
    current_ids = set()
    for rec in current:
        current_ids.add(rec["id"])
        old = previous.get(rec["id"])
        if old is None:
            changes.append({"type": "added", "register": register, "id": rec["id"], "name": rec.get("name")})
        elif _comparable(old) != _comparable(rec):
            entry = {"type": "changed", "register": register, "id": rec["id"], "name": rec.get("name")}
            detail = describe_record_change(old, rec)
            if detail:
                entry["detail"] = detail
            changes.append(entry)
    for old_id, old_rec in previous.items():
        if old_id not in current_ids:
            entry = {"type": "removed", "register": register, "id": old_id, "name": old_rec.get("name")}
            date_field = REGISTRATION_DATE_FIELD.get(register)
            if date_field and old_rec.get(date_field):
                entry["registered_on"] = old_rec[date_field]
            changes.append(entry)
    return changes


def run(fetcher: Callable[[str], list[dict]] = fetch_csv) -> int:
    """Runs one full scrape cycle. `fetcher` is injectable for tests."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)

    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    all_changes: list[dict] = []
    counts: dict[str, int] = {}
    errors: dict[str, str] = {}

    for register, code in SOURCES.items():
        try:
            rows = fetcher(code)
        except Exception as exc:  # one register failing shouldn't kill the whole run
            errors[register] = str(exc)
            print(f"[warn] failed to fetch {register} ({code}): {exc}", file=sys.stderr)
            continue

        previous = load_previous(register)
        current = NORMALIZERS[register](rows)

        if register == "casps":
            # AFM's own crypto register tends to list a Dutch CASP's new
            # service/authorisation before ESMA's slower, EU-wide export
            # does - fold it in here rather than as a separate register
            # (see merge_esma_and_afm_casps). AFM being unreachable
            # shouldn't fail the run - just fall back to ESMA-only data.
            try:
                afm_rows = fetch_afm_rows()
                afm_current = normalize_afm(afm_rows)
                current = merge_esma_and_afm_casps(current, afm_current, previous)
            except Exception as exc:
                print(f"[warn] failed to fetch/merge AFM register: {exc}", file=sys.stderr)
                current = [{**rec, "source": "esma"} for rec in current]

        changes = diff_records(register, previous, current)
        all_changes.extend(changes)
        counts[register] = len(current)

        out_path = DATA_DIR / f"{register}.json"
        with out_path.open("w", encoding="utf-8") as f:
            json.dump(
                {"register": register, "generated_at": now, "records": current},
                f, ensure_ascii=False, indent=2,
            )

        if changes:
            print(f"[info] {register}: {len(changes)} change(s), {len(current)} total record(s)")
        else:
            print(f"[info] {register}: no change ({len(current)} records)")

    meta_path = DATA_DIR / "meta.json"
    meta = {}
    if meta_path.exists():
        with meta_path.open(encoding="utf-8") as f:
            meta = json.load(f)
    meta["last_checked"] = now
    meta["record_counts"] = counts
    if errors:
        meta["last_errors"] = errors
    else:
        meta.pop("last_errors", None)
    if all_changes:
        meta["last_change_at"] = now
    with meta_path.open("w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

    changelog_path = HISTORY_DIR / "changelog.json"
    changelog = []
    if changelog_path.exists():
        with changelog_path.open(encoding="utf-8") as f:
            changelog = json.load(f)
    for change in all_changes:
        changelog.append({**change, "timestamp": now})
    with changelog_path.open("w", encoding="utf-8") as f:
        json.dump(changelog, f, ensure_ascii=False, indent=2)

    print(f"[info] done - {len(all_changes)} total change(s) across all registers")

    # Every run rewrites each register's `generated_at` to "now", so a git
    # diff on data/ is never empty even when ESMA's actual content didn't
    # change - that's just freshness bookkeeping, not "new data". Expose the
    # real, record-level change count (and a human-readable per-entity
    # breakdown: which party, which register, added/changed/removed) as a
    # GitHub Actions step output, so the workflow can both gate its Slack
    # notification on genuinely new results instead of every successful run,
    # and say *what* changed rather than just how many.
    github_output = os.environ.get("GITHUB_OUTPUT")
    if github_output:
        # Emoji type-indicator instead of a text label ("toegevoegd"/
        # "gewijzigd"/"verwijderd") - faster to scan a list of these than a
        # column of words, and it means the entity name can lead the line
        # instead of being buried after "register: ".
        TYPE_EMOJI = {"added": "🆕", "changed": "✏️", "removed": "❌"}
        MAX_SUMMARY_LINES = 20
        detail_lines = []
        for c in sorted(all_changes, key=_change_priority_key):
            emoji = TYPE_EMOJI.get(c["type"], "•")
            register_label = REGISTER_LABELS_NL.get(c["register"], c["register"])
            name = c.get("name") or c["id"]
            # One line per changed *record*, not per changed aspect -
            # summarize_change_detail() folds every item in c["detail"] (e.g.
            # 7 services all gaining the same new country) into a single
            # combined, capped segment string, instead of the old behaviour
            # of repeating the entity's own name once per detail item.
            if c["type"] == "changed" and c.get("detail"):
                detail_lines.append(f"{emoji} {name} _({register_label})_ — {summarize_change_detail(c['detail'])}")
            else:
                detail_lines.append(f"{emoji} {name} _({register_label})_")
        summary_lines = detail_lines[:MAX_SUMMARY_LINES]
        if len(detail_lines) > MAX_SUMMARY_LINES:
            summary_lines.append(f"... en {len(detail_lines) - MAX_SUMMARY_LINES} andere wijziging(en)")
        with open(github_output, "a", encoding="utf-8") as f:
            f.write(f"real_changes={len(all_changes)}\n")
            f.write("change_summary<<EOF\n")
            f.write("\n".join(summary_lines) + "\n")
            f.write("EOF\n")

    if len(errors) == len(SOURCES):
        return 1  # every single register failed - a real failure, not "no change"
    return 0


def main() -> None:
    sys.exit(run())


if __name__ == "__main__":
    main()
